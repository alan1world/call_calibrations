#! /usr/bin/env python
#! -*- coding: utf-8 -*-

import itertools

from PyQt5 import QtWidgets, QtGui, QtCore

from openpyxl import Workbook
import openpyxl.styles

from calibrations_sql import CalibrationStore

class CalibrationExportSummary(QtWidgets.QDialog):
    
    def __init__(self, parent=None):

        super().__init__(parent=parent)
        self.output = dict()
        self.initUI()

    def initUI(self):
        hbox_main_panel = QtWidgets.QHBoxLayout()
        self.setLayout(hbox_main_panel)

        vbox_button_panel = QtWidgets.QVBoxLayout()
        save_button = QtWidgets.QPushButton('Export', self)
        cancel_button = QtWidgets.QPushButton('Close', self)
        vbox_button_panel.addWidget(save_button)
        vbox_button_panel.addWidget(cancel_button)
        vbox_button_panel.setAlignment(save_button,QtCore.Qt.AlignTop)
        vbox_button_panel.addStretch()

        self.listWidget = QtWidgets.QListWidget()
        years = ('2019', '2018')
        #months = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 
        #          'Oct', 'Nov', 'Dec')
        #for element in itertools.product('2018','Jan'):
        for year in years:
            for month in range(1,53):
            #for month in months:
                submit = year + ' - Week ' + str(month)
                self.listWidget.addItem(submit)

        hbox_main_panel.addWidget(self.listWidget)
        hbox_main_panel.addLayout(vbox_button_panel)
        self.setWindowTitle('Export Call Calibrations')

        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        save_button.clicked.connect(self.export_window)
        cancel_button.clicked.connect(self.call_cancel)


    def export_window(self):
        ev = CalibrationStore()
        wb = Workbook()
        ws = wb.active
        collison_colours = {'white':'ffffff', 'magenta':'CE0058', 'red':'96172E', 'blue':'007DBA', 'dark_blue':'003865',
                            'pink':'E89CAE', 'teal':'7CE0D3', 'grey':'4B4F54'}
        collison_font = openpyxl.styles.Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='ffffff')
        collison_fill = openpyxl.styles.PatternFill(fill_type=None,start_color=collison_colours['magenta'],end_color=collison_colours['magenta'])
        fixed_format_cells = (('A1','Year:'),('A2','Week:'),('A3','Score:'),('A4','InteractionFlow:'),('A5','FirstcontactResolution:'),
                              ('A6','Communication:'),('A7','CustomerFocus:'),('A8','Demeanor:'))
        for loc in fixed_format_cells:
            ws[loc[0]] = loc[1]
            ws[loc[0]].font = collison_font
            ws[loc[0]].fill = collison_fill
        for position, item in enumerate(self.listWidget.selectedItems(),2):
            first, second = item.text().split(' - Week ')
            ws.cell(row=1,column=position, value=int(first))
            ws.cell(row=2,column=position, value=int(second))
            ws.cell(row=1,column=position).font = collison_font
            ws.cell(row=2,column=position).font = collison_font
            ws.cell(row=1,column=position).fill = collison_fill
            ws.cell(row=2,column=position).fill = collison_fill
            print(f'Year: {first}, Week: {second}')
            export_answer = ev.export_weekly_score(year_in=first, week_in=int(second))
            if export_answer[0] is not None:
                ws.cell(row=3,column=position, value=int(export_answer[0]))
                ws.cell(row=4,column=position, value=int(export_answer[1]))
                ws.cell(row=5,column=position, value=int(export_answer[2]))
                ws.cell(row=6,column=position, value=int(export_answer[3]))
                ws.cell(row=7,column=position, value=int(export_answer[4]))
                ws.cell(row=8,column=position, value=int(export_answer[5]))
                print(f"Score: {export_answer[0]:.1f}\n"
                    f"InteractionFlow: {export_answer[1]:.1f}\n"
                    f"FirstcontactResolution: {export_answer[2]:.1f}\n"
                    f"Communication: {export_answer[3]:.1f}\n"
                    f"CustomerFocus: {export_answer[4]:.1f}\n"
                    f"Demeanor: {export_answer[5]:.1f}")
        # Get list of selected  year / weeks
        # Send to database each year-week
        # Export list of results to output file
        wb.save('export.xlsx')
        self.accept()
        
    def call_cancel(self):
        self.close()

    @staticmethod
    def getCalibrationDetails(parent = None):
        dialog = CalibrationExportSummary(parent)
        result = dialog.exec_()
        if result == QtWidgets.QDialog.Accepted:
            return_result = dict()
            return_result = dialog.output
            return (return_result, True)
        else:
            return (dict(), False)